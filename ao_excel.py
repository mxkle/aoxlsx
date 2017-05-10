from openpyxl import load_workbook
from collections import OrderedDict

def init_wb(i):
        wb = load_workbook('excel.xlsx')
        try:
                ws = wb[i]
        except:
                ws = wb.create_sheet(i)
                hl = ['Date','Action','Amount','Price/1','Price/total','Tier','Location']
                for col in ws.iter_cols(min_row=1, max_col=7, max_row=1):
                        for cell in col:
                                cell.value=hl.pop(0)

        return (ws,wb)

def get_max(ws):
        max = ws.max_row
        c = 'A'+str(max)
        while ws[c].value == None:
                max -= 1
                print(ws[c].value)
                c = 'A'+str(max)

        return(max)

def insert_values(ws,vals):
        max = get_max(ws)
        print(max)
        if ws['A'+str(max)].value != vals['date']: 
                max += 1
                ws.append(list(vals.values()))
        
        else:
                for col in ws.iter_cols(min_row=1, max_col=7, max_row=1):
                        for cell in col:
                                #print(cell.value)
                                if cell.value == "Amount":
                                        i=int(cell.offset(row=max-1).value)
                                        i+=vals['a']
                                        cell.offset(row=max-1).value=i
                                if cell.value == "Price/total":
                                        i=int(cell.offset(row=max-1).value)
                                        i+=vals['P']
                                        cell.offset(row=max-1).value=i
                                    
def save_file(wb):
        wb.save(filename = "excel.xlsx")


